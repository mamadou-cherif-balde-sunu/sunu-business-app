from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, FileResponse
from django.conf import settings
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

from .models import ReponseClient, DemandeModification
from .utils_report import get_week_range, get_month_range, label_week, label_month, get_period_stats
from .pptx_report import generate_report

# ==========================================
# 1. AUTHENTIFICATION
# ==========================================

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Identifiant ou mot de passe incorrect.')
    return render(request, 'relation_client/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')


# ==========================================
# 2. ACCUEIL ET RAPPORTS
# ==========================================

@login_required(login_url='login')
def dashboard(request):
    reponses = ReponseClient.objects.all()
    total = reponses.count()

    def calc(qs):
        t = qs.count()
        if t == 0:
            return {'nps': 0, 'csat': 0, 'ces': 0, 'total': 0}
        promoteurs  = qs.filter(categorie_nps='Promoteur').count()
        detracteurs = qs.filter(categorie_nps='Detracteur').count()
        nps  = round((promoteurs/t*100) - (detracteurs/t*100), 1)
        csat = round(qs.filter(csat='Satisfait').count()/t*100, 1)
        eleve  = qs.filter(ces='Effort eleve').count()
        faible = qs.filter(ces='Effort faible').count()
        ces = round((eleve/t*100) - (faible/t*100), 1)
        return {'nps': nps, 'csat': csat, 'ces': ces, 'total': t}

    context = {
        'global_stats': calc(reponses),
        'vie_stats':    calc(reponses.filter(sous_canal__icontains='VIE')),
        'iard_stats':   calc(reponses.filter(sous_canal__icontains='IARD')),
        'total':        total,
        'joints':       reponses.filter(statut_appel='Joint').count(),
        'injoignables': reponses.filter(statut_appel='Injoignable').count(),
    }
    return render(request, 'relation_client/dashboard.html', context)


@login_required(login_url='login')
def indicateurs(request):
    reponses = ReponseClient.objects.all()

    def calc(qs):
        t = qs.count()
        if t == 0:
            return None
        promoteurs  = qs.filter(categorie_nps='Promoteur').count()
        detracteurs = qs.filter(categorie_nps='Detracteur').count()
        passifs     = qs.filter(categorie_nps='Passif').count()
        nps  = round((promoteurs/t*100) - (detracteurs/t*100), 1)
        csat = round(qs.filter(csat='Satisfait').count()/t*100, 1)
        eleve  = qs.filter(ces='Effort eleve').count()
        faible = qs.filter(ces='Effort faible').count()
        ces = round((eleve/t*100) - (faible/t*100), 1)
        return {
            'nps': nps, 'csat': csat, 'ces': ces, 'total': t,
            'promoteurs': promoteurs, 'detracteurs': detracteurs,
            'passifs': passifs,
            'joints':      qs.filter(statut_appel='Joint').count(),
            'injoignables':qs.filter(statut_appel='Injoignable').count(),
            'pas_reponse': qs.filter(statut_appel='Pas de reponse').count(),
        }

    tableau = [
        {'perimetre': 'VIE',    'stats': calc(reponses.filter(sous_canal__icontains='VIE'))},
        {'perimetre': 'IARD',   'stats': calc(reponses.filter(sous_canal__icontains='IARD'))},
        {'perimetre': 'GLOBAL', 'stats': calc(reponses)},
    ]
    return render(request, 'relation_client/indicateurs.html', {'tableau': tableau})


# ==========================================
# 3. SAISIE ET CONSULTATION
# ==========================================

@login_required(login_url='login')
def saisie(request):
    if request.method == 'POST':
        try:
            nps_intervalle = request.POST.get('nps_score', '').strip()
            nps_score = None
            categorie_nps = request.POST.get('categorie_nps', '')

            if nps_intervalle == "10":
                nps_score = 10
            elif nps_intervalle == "8":
                nps_score = 8
            elif nps_intervalle == "6":
                nps_score = 6

            # Récupération de l'objet de visite ou de l'alternative précisée
            objet_visite = request.POST.get('objet_visite')
            if objet_visite == "Autre":
                objet_visite = request.POST.get('objet_visite_autre')

            ReponseClient.objects.create(
                cc                   = request.user,
                date_enregistrement  = request.POST.get('date_enregistrement') or date.today(),
                nom_prenoms          = request.POST.get('nom_prenoms'),
                contact              = request.POST.get('contact'),
                numero_police        = request.POST.get('numero_police'),
                objet_visite         = objet_visite,
                canal                = request.POST.get('canal'),
                sous_canal           = request.POST.get('sous_canal'), # Ajout du Périmètre choisi
                date_appel           = request.POST.get('date_appel') or date.today(),
                statut_appel         = request.POST.get('statut_appel'),
                nps_score            = nps_score,
                categorie_nps        = categorie_nps,
                csat                 = request.POST.get('csat', ''),
                ces                  = request.POST.get('ces', ''),
                motif_insatisfaction = request.POST.get('motif_insatisfaction'),
                observations         = request.POST.get('observations'),
            )
            messages.success(request, 'Réponse enregistrée avec succès ✅')
            return redirect('saisie')
        except Exception as e:
            messages.error(request, f'Erreur : {e}')
    return render(request, 'relation_client/saisie.html')


@login_required(login_url='login')
def liste(request):
    reponses = ReponseClient.objects.all().order_by('-date_appel')
    
    # Récupération des filtres depuis le template de l'application
    canal      = request.GET.get('canal')
    sous_canal = request.GET.get('sous_canal')
    indicateur = request.GET.get('indicateur')
    statut     = request.GET.get('statut')
    nom_client = request.GET.get('nom_client', '').strip()
    tel_client = request.GET.get('tel_client', '').strip()
    date_debut = request.GET.get('date_debut')
    date_fin   = request.GET.get('date_fin')

    # Application des filtres cumulés
    if canal: 
        reponses = reponses.filter(canal=canal)
    
    if sous_canal:
        reponses = reponses.filter(sous_canal=sous_canal)
        
    if indicateur:
        if indicateur == 'NPS':
            reponses = reponses.exclude(categorie_nps='')
        elif indicateur == 'CSAT':
            reponses = reponses.filter(csat='Satisfait')
        elif indicateur == 'CES':
            reponses = reponses.filter(ces__in=['Effort eleve', 'Effort faible'])

    if statut: 
        reponses = reponses.filter(statut_appel=statut)
        
    if nom_client:
        reponses = reponses.filter(nom_prenoms__icontains=nom_client)
        
    if tel_client:
        reponses = reponses.filter(contact__icontains=tel_client)
        
    if date_debut: 
        reponses = reponses.filter(date_appel__gte=date_debut)
        
    if date_fin:   
        reponses = reponses.filter(date_appel__lte=date_fin)

    return render(request, 'relation_client/liste.html', {'reponses': reponses})

@login_required(login_url='login')
def profil(request):
    mes_reponses = ReponseClient.objects.filter(cc=request.user).order_by('-date_appel')
    total = mes_reponses.count()

    def calc(qs):
        t = qs.count()
        if t == 0:
            return {'nps': 0, 'csat': 0, 'ces': 0, 'total': 0}
        promoteurs  = qs.filter(categorie_nps='Promoteur').count()
        detracteurs = qs.filter(categorie_nps='Detracteur').count()
        nps  = round((promoteurs/t*100) - (detracteurs/t*100), 1)
        csat = round(qs.filter(csat='Satisfait').count()/t*100, 1)
        eleve  = qs.filter(ces='Effort eleve').count()
        faible = qs.filter(ces='Effort faible').count()
        ces = round((eleve/t*100) - (faible/t*100), 1)
        return {'nps': nps, 'csat': csat, 'ces': ces, 'total': t}

    stats = calc(mes_reponses)
    context = {
        'mes_reponses': mes_reponses,
        'total': total,
        'stats': stats,
    }
    return render(request, 'relation_client/profil.html', context)


# ==========================================
# 4. EXPORTS ET EXTRACTIONS
# ==========================================

@login_required(login_url='login')
def export(request):
    reponses = ReponseClient.objects.all().order_by('-date_appel')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reponses Clients"
    entetes = [
        'Canal', 'Périmètre', 'Date Enregistrement', 'CC',
        'Nom et Prenoms', 'Contact', 'N Police',
        'Objet Visite', 'Date Appel', 'Statut Appel',
        'Score NPS', 'Categorie NPS', 'CSAT', 'CES',
        'Motif Insatisfaction', 'Observations'
    ]
    for col, entete in enumerate(entetes, start=1):
        cell = ws.cell(row=1, column=col, value=entete)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='E30613')
        cell.alignment = Alignment(horizontal='center')
        
    for row, r in enumerate(reponses, start=2):
        ws.cell(row=row, column=1,  value=r.canal)
        ws.cell(row=row, column=2,  value=r.sous_canal) # Intégration du Périmètre dans l'export Excel
        ws.cell(row=row, column=3,  value=str(r.date_enregistrement))
        ws.cell(row=row, column=4,  value=r.cc.username)
        ws.cell(row=row, column=5,  value=r.nom_prenoms)
        ws.cell(row=row, column=6,  value=r.contact)
        ws.cell(row=row, column=7,  value=r.numero_police)
        ws.cell(row=row, column=8,  value=r.objet_visite)
        ws.cell(row=row, column=9,  value=str(r.date_appel))
        ws.cell(row=row, column=10, value=r.statut_appel)
        ws.cell(row=row, column=11, value=r.nps_score)
        ws.cell(row=row, column=12, value=r.categorie_nps)
        ws.cell(row=row, column=13, value=r.csat)
        ws.cell(row=row, column=14, value=r.ces)
        ws.cell(row=row, column=15, value=r.motif_insatisfaction)
        ws.cell(row=row, column=16, value=r.observations)
        
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="SUNU_Reponses.xlsx"'
    wb.save(response)
    return response


def export_pptx(request):
    mode = request.GET.get("mode", "semaine")

    if mode == "mois":
        annee = int(request.GET.get("annee"))
        mois = int(request.GET.get("mois"))
        debut, fin = get_month_range(annee, mois)
        period_label = label_month(annee, mois)
        filename = f"Rapport_{label_month(annee, mois).replace(' ', '_')}.pptx"
    else:
        jour_str = request.GET.get("date")
        jour = date.fromisoformat(jour_str) if jour_str else date.today()
        debut, fin = get_week_range(jour)
        period_label = label_week(debut, fin)
        filename = f"Rapport_Semaine_{debut.strftime('%d-%m')}_au_{fin.strftime('%d-%m-%Y')}.pptx"

    stats = get_period_stats(ReponseClient, debut, fin)

    output_dir = os.path.join(settings.BASE_DIR, "media", "rapports_pptx")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, filename)

    generate_report(period_label, stats, output_path)

    return FileResponse(open(output_path, "rb"), as_attachment=True, filename=filename)


# ==========================================
# 5. WORKFLOW MODIFICATIONS ET AGENTS
# ==========================================

@login_required(login_url='login')
def demande_modification(request, saisie_id):
    saisie = get_object_or_404(ReponseClient, id=saisie_id, cc=request.user)
    if request.method == 'POST':
        champs          = request.POST.getlist('champ[]')
        nouvelles_valeurs = request.POST.getlist('nouvelle_valeur[]')
        motif_global    = request.POST.get('motif_global', '')

        for champ, valeur in zip(champs, nouvelles_valeurs):
            if champ and valeur:
                DemandeModification.objects.create(
                    saisie          = saisie,
                    demandeur       = request.user,
                    champ           = champ,
                    nouvelle_valeur = valeur,
                    motif           = motif_global,
                )

        messages.success(request, 'Demande(s) envoyée(s) à l\'administrateur ✅')
        return redirect('profil')
    return render(request, 'relation_client/demande_modification.html', {'saisie': saisie})


# ==========================================
# 6. ESPACE ADMINISTRATION (STAFF)
# ==========================================

@staff_member_required(login_url='login')
def gestion_utilisateurs(request):
    if request.method == 'POST':
        username   = request.POST.get('username')
        password   = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name  = request.POST.get('last_name')
        email      = request.POST.get('email')
        is_staff   = request.POST.get('is_staff') == 'on'
        if User.objects.filter(username=username).exists():
            messages.error(request, f'Le nom utilisateur {username} existe déjà.')
        else:
            User.objects.create_user(
                username=username, password=password,
                first_name=first_name, last_name=last_name,
                email=email, is_staff=is_staff
            )
            messages.success(request, f'Compte {username} créé avec succès ✅')
    utilisateurs = User.objects.all().order_by('-date_joined')
    return render(request, 'relation_client/utilisateurs.html', {'utilisateurs': utilisateurs})


@staff_member_required(login_url='login')
def toggle_utilisateur(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if user != request.user:
        user.is_active = not user.is_active
        user.save()
        statut = 'activé' if user.is_active else 'désactivé'
        messages.success(request, f'Compte {user.username} {statut} ✅')
    return redirect('gestion_utilisateurs')


@staff_member_required(login_url='login')
def modifier_saisie(request, saisie_id):
    saisie = get_object_or_404(ReponseClient, id=saisie_id)
    if request.method == 'POST':
        try:
            nps_score = request.POST.get('nps_score')
            nps_score = int(nps_score) if nps_score else None
            saisie.date_enregistrement  = request.POST.get('date_enregistrement') or saisie.date_enregistrement
            saisie.nom_prenoms          = request.POST.get('nom_prenoms')
            saisie.contact              = request.POST.get('contact')
            saisie.numero_police        = request.POST.get('numero_police')
            saisie.objet_visite         = request.POST.get('objet_visite')
            saisie.canal                = request.POST.get('canal')
            saisie.sous_canal           = request.POST.get('sous_canal') # Ajout modifiable du Périmètre
            saisie.date_appel           = request.POST.get('date_appel') or saisie.date_appel
            saisie.statut_appel         = request.POST.get('statut_appel')
            saisie.nps_score            = nps_score
            saisie.categorie_nps        = request.POST.get('categorie_nps', '')
            saisie.csat                 = request.POST.get('csat', '')
            saisie.ces                  = request.POST.get('ces', '')
            saisie.motif_insatisfaction = request.POST.get('motif_insatisfaction')
            saisie.observations         = request.POST.get('observations')
            saisie.save()
            messages.success(request, 'Saisie modifiée avec succès ✅')
            return redirect('liste')
        except Exception as e:
            messages.error(request, f'Erreur : {e}')
    return render(request, 'relation_client/modifier_saisie.html', {'saisie': saisie})


@staff_member_required(login_url='login')
def demandes_admin(request):
    statut_filtre = request.GET.get('statut', '')
    demandes = DemandeModification.objects.all().order_by('-date_demande')
    if statut_filtre:
        demandes = demandes.filter(statut=statut_filtre)

    demandes_attente = DemandeModification.objects.filter(statut='En attente').count()

    if request.method == 'POST':
        demande_id = request.POST.get('demande_id')
        action     = request.POST.get('action')
        demande    = get_object_or_404(DemandeModification, id=demande_id)

        if action == 'approuver':
            try:
                saisie = demande.saisie
                champ  = demande.champ
                valeur = demande.nouvelle_valeur

                if champ == 'nps_score':
                    valeur = int(valeur)
                    saisie.nps_score = valeur
                    if valeur >= 9:
                        saisie.categorie_nps = 'Promoteur'
                    elif valeur >= 7:
                        saisie.categorie_nps = 'Passif'
                    else:
                        saisie.categorie_nps = 'Detracteur'
                elif champ == 'categorie_nps':
                    saisie.categorie_nps = valeur
                elif champ == 'csat':
                    saisie.csat = valeur
                elif champ == 'ces':
                    saisie.ces = valeur
                elif champ == 'statut_appel':
                    saisie.statut_appel = valeur
                elif champ == 'canal':
                    saisie.canal = valeur
                elif champ == 'sous_canal': # Géré au cas où l'admin approuve ce champ spécifique
                    saisie.sous_canal = valeur
                elif champ == 'nom_prenoms':
                    saisie.nom_prenoms = valeur
                elif champ == 'contact':
                    saisie.contact = valeur
                elif champ == 'observations':
                    saisie.observations = valeur

                saisie.save()
                demande.statut = 'Approuvee'
                demande.save()
                messages.success(request, f'Modification de "{champ}" approuvée ✅')

            except Exception as e:
                messages.error(request, f'Erreur : {e}')

        elif action == 'rejeter':
            demande.statut = 'Rejetee'
            demande.save()
            messages.warning(request, 'Demande rejetée.')

        return redirect('demandes_admin')

    return render(request, 'relation_client/demandes_admin.html', {
        'demandes': demandes,
        'demandes_attente': demandes_attente,
        'statut_filtre': statut_filtre,
    })